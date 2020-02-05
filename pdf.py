from trace import Trace

from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO
from openpyxl import Workbook
from openpyxl import load_workbook


def convertPdf(file):
    resourceManager = PDFResourceManager()
    _string = StringIO()
    # codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(resourceManager, _string, laparams=laparams)
    fp = open(file, 'rb')
    interpreter = PDFPageInterpreter(resourceManager, device)
    password = ""
    maxpages = 0
    catching = True
    pagenos = set()
    # 가져오는거 1번쨰 두번쨰 사이에 엔터(NULL)
    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password, caching=catching,
                                  check_extractable=True):
        interpreter.process_page(page)
    text = _string.getvalue()
    temp = text.split("\n")
    del temp[-1]
    del temp[-1]
    fp.close()
    device.close()
    _string.close()
    return temp


# file = 'C:\\Users\\HS YUN\\Desktop\\pythonpdf\\1. pdf_to_text.pdf'
# print(convertPdf(file))

def createExcel():
    write_wb = Workbook()
    # 이름이 있는 시트를 생성
    write_ws = write_wb.create_sheet('title')

    # Sheet1 에다 입력
    write_ws = write_wb.active
    write_ws['A1'] = '숫자'
    # 행 단위로 추가
    write_ws.append([1, 2, 3])

    # 셀 단위로 추가
    write_ws.cell(5, 5, '5 행 5 열')
    filePath = 'C:\\Users\\HS YUN\\Desktop\\pythonpdf\\test.xlsx'
    write_wb.save(filePath)


# createExcel()

def readExcel():
    # data_only Ture 로 해줘야 수식이 아닌 값으로 받아온다
    filePath = 'C:\\Users\\HS YUN\\Desktop\\pythonpdf\\test.xlsx'
    load_wb = load_workbook(filePath, data_only=True)

    # 시트 이름으로 불러오기
    load_ws = load_wb['Sheet']

    # 셀 주소로 값 출력
    print(load_ws['A1'].value)
    # 셀 좌표로 값 출력
    print(load_ws.cell(1, 2).value)

    # 지정한 셀 출력
    get_cells = load_ws['A1':'D2']
    for row in get_cells:
        for cell in row:
            print(cell.value)

    # 모든 행 단위로 출력
    load_ws = load_wb['Sheet']
    for row in load_ws.rows:
        print(row)

    # 모든 열 단위로 출력
    load_ws = load_wb['Sheet']
    for column in load_ws.columns:
        print(column)

    # 모든 행과 열 출력
    load_ws = load_wb['Sheet']
    allVal = []
    for row in load_ws.rows:
        rowVal = []
        for cell in row:
            rowVal.append(cell.value)
        allVal.append(rowVal)
    print(allVal)

# readExcel()


def exercise():
    pdfPath = 'C:\\Users\\HS YUN\\Desktop\\pythonpdf\\2. pdf_to_excel.pdf'
    text = convertPdf(pdfPath)
    print(text)
    name = []
    physics = []
    math = []
    for i in text:
        name.append(i[3:6])
        physics.append(i[12:14])
        math.append(i[20:22])
    # print(name)
    # print(physics)
    # print(math)
    physics = list(map(int, physics))
    math = list(map(int, math))
    sum = []
    for i in range(4):
        sum.append(int(math[i] + physics[i]))
    avg = map(lambda x: x/2, sum)
    avg = list(map(float, avg))
    print(avg)
    letter = []
    for i in avg:
        if i>=90:
            val = 'A+'
        elif i>=80:
            val = 'A'
        elif i>=70:
            val = 'B+'
        elif i>=60:
            val = 'B'
        elif i>=50:
            val = 'C+'
        elif i>=40:
            val = 'C'
        else:
            val = 'F'
        letter.append(val)

    for i in range(4):
        print('name :{}, physics :{}, math :{}, sum :{}, avg :{}, letter :{}'.format(name[i], physics[i], math[i], sum[i], avg[i], letter[i]))

    write_wb = Workbook()
    # 이름이 있는 시트를 생성
    write_ws = write_wb.create_sheet('title')

    # Sheet1 에다 입력
    write_ws = write_wb.active
    write_ws['A1'] = '중간 성적'

    # 행 단위로 추가
    write_ws.append(['이름', '물리', '미적', '총합', '평균', '학점'])

    # 셀 단위로 추가
    arr = name + physics + math + sum + avg + letter
    print(arr)
    for i in range(0, 4):
        for j in range(0, 6):
            write_ws.cell(i+3, j+1, arr[i + 4*j])
    filePath = 'C:\\Users\\HS YUN\\Desktop\\pythonpdf\\test2.xlsx'
    write_wb.save(filePath)


# exercise()